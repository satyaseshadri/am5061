"""AM5061 - Design of Thermal and Fluid Systems.
Shared helpers for the course notebooks.

Design of Thermal and Fluid Systems, IIT Madras, Jul-Nov 2026.

This module is deliberately thin. It wraps CoolProp with names and units that
match the lecture notation, adds an Excel writer that produces workbooks you
can actually filter, and sets a consistent plot style. It does NOT hide the
engineering: every case notebook still writes its own equations.

Install (first cell of any Colab notebook):
    !pip install -q CoolProp openpyxl

Units are SI throughout, with ONE exception that is flagged everywhere it
appears: temperatures in function arguments named `..._C` are in Celsius,
because that is how the case briefs state them. Everything internal is kelvin.
"""
from __future__ import annotations

import math

from CoolProp.CoolProp import PropsSI, PhaseSI

__all__ = [
    "K", "C", "State", "state", "sat_liquid", "sat_vapour", "p_sat", "T_sat",
    "glide", "h_fg", "critical", "fluids", "solve", "sweep", "to_excel",
    "style_plots", "NAVY", "ORANGE", "BLUE", "MUTED",
    "lmtd", "effectiveness", "ntu_required", "exergy", "T0_REF", "P0_REF",
]

# ---------------------------------------------------------------- constants
NAVY, ORANGE, BLUE, MUTED = "#1F3864", "#ED7D31", "#4472C4", "#59626E"
T0 = 273.15


def K(t_celsius: float) -> float:
    """Celsius -> kelvin. Use this at the boundary, never inside a formula."""
    return t_celsius + T0


def C(t_kelvin: float) -> float:
    """Kelvin -> Celsius, for reporting only."""
    return t_kelvin - T0


# ------------------------------------------------------------------- states
class State:
    """A thermodynamic state. Immutable, and it knows its own fluid.

    Construct it with any two independent properties:
        State("R134a", P=1e6, T=K(70))
        State("Water", P=101325, Q=0)      # saturated liquid
        State("R134a", P=p_cond, H=h2)

    Then read properties as attributes: .T .p .h .s .d .cp .x
    Attribute names match the lecture notation, not CoolProp's letter codes,
    so a student reading the notebook does not need the CoolProp manual open.
    """

    _MAP = {"T": "T", "P": "P", "H": "H", "S": "S", "D": "D", "Q": "Q"}

    def __init__(self, fluid: str, **kw):
        if len(kw) != 2:
            raise ValueError(
                f"a state needs exactly two properties, got {list(kw)}. "
                "Two and only two - that is the phase rule, not a quirk."
            )
        (n1, v1), (n2, v2) = kw.items()
        for n in (n1, n2):
            if n not in self._MAP:
                raise ValueError(f"unknown property {n!r}; use T, P, H, S, D or Q")
        self.fluid, self._args = fluid, (n1, v1, n2, v2)

    def _get(self, what: str) -> float:
        n1, v1, n2, v2 = self._args
        return PropsSI(what, n1, v1, n2, v2, self.fluid)

    # Named so they read like the equations on the slides.
    T  = property(lambda s: s._get("T"),  doc="temperature, K")
    p  = property(lambda s: s._get("P"),  doc="pressure, Pa")
    h  = property(lambda s: s._get("H"),  doc="specific enthalpy, J/kg")
    s  = property(lambda s: s._get("S"),  doc="specific entropy, J/kg.K")
    d  = property(lambda s: s._get("D"),  doc="density, kg/m3")
    cp = property(lambda s: s._get("C"),  doc="cp, J/kg.K")
    mu = property(lambda s: s._get("V"),  doc="dynamic viscosity, Pa.s")
    k  = property(lambda s: s._get("L"),  doc="thermal conductivity, W/m.K")
    x  = property(lambda s: s._get("Q"),  doc="vapour quality, - (=-1 if single phase)")

    @property
    def T_C(self) -> float:
        return C(self.T)

    @property
    def phase(self) -> str:
        n1, v1, n2, v2 = self._args
        return PhaseSI(n1, v1, n2, v2, self.fluid)

    def __repr__(self):
        try:
            return (f"State({self.fluid}: {self.T_C:.2f} C, {self.p/1e5:.3f} bar, "
                    f"h={self.h/1e3:.2f} kJ/kg, {self.phase})")
        except Exception:
            return f"State({self.fluid}, {self._args})"


def state(fluid: str, **kw) -> State:
    """Shorthand for State(...)."""
    return State(fluid, **kw)


def sat_liquid(fluid: str, *, T=None, p=None) -> State:
    """Saturated liquid at T or p. Give one, not both."""
    if (T is None) == (p is None):
        raise ValueError("give exactly one of T or p")
    return State(fluid, T=T, Q=0) if T is not None else State(fluid, P=p, Q=0)


def sat_vapour(fluid: str, *, T=None, p=None) -> State:
    """Saturated vapour at T or p."""
    if (T is None) == (p is None):
        raise ValueError("give exactly one of T or p")
    return State(fluid, T=T, Q=1) if T is not None else State(fluid, P=p, Q=1)


def p_sat(fluid: str, T: float) -> float:
    """Saturation pressure, Pa. For a BLEND this is the bubble-point pressure."""
    return PropsSI("P", "T", T, "Q", 0, fluid)


def T_sat(fluid: str, p: float) -> float:
    """Saturation temperature, K.

    WARNING for blends: a zeotropic mixture has no single saturation
    temperature. This returns the BUBBLE point. Use glide() to see the spread.
    """
    return PropsSI("T", "P", p, "Q", 0, fluid)


def glide(fluid: str, p: float) -> float:
    """Dew minus bubble temperature at p, K. Zero for a pure fluid."""
    return (PropsSI("T", "P", p, "Q", 1, fluid)
            - PropsSI("T", "P", p, "Q", 0, fluid))


def h_fg(fluid: str, *, T=None, p=None) -> float:
    """Latent heat, J/kg."""
    return sat_vapour(fluid, T=T, p=p).h - sat_liquid(fluid, T=T, p=p).h


def critical(fluid: str) -> dict:
    """Critical point, for checking you are not extrapolating past it."""
    return {"T": PropsSI("TCRIT", fluid), "p": PropsSI("PCRIT", fluid)}


def fluids() -> list:
    """Every fluid CoolProp knows. There are about 130."""
    import CoolProp
    return sorted(CoolProp.__fluids__)


# ------------------------------------------------------------------ solvers
def solve(f, x0, *, tol=1e-10, max_iter=200, bracket=None):
    """Find x where f(x) = 0.

    Uses Brent's method when you give a bracket (robust, always converges if
    the bracket is valid), otherwise secant from x0. Raises with a readable
    message rather than returning a wrong answer silently, which is the whole
    problem with doing this in a spreadsheet.
    """
    from scipy.optimize import brentq, newton
    if bracket is not None:
        a, b = bracket
        fa, fb = f(a), f(b)
        if fa * fb > 0:
            raise ValueError(
                f"f({a:g})={fa:g} and f({b:g})={fb:g} have the same sign, so no "
                "root is bracketed. Widen the bracket or check the equation."
            )
        return brentq(f, a, b, xtol=tol, maxiter=max_iter)
    return newton(f, x0, tol=tol, maxiter=max_iter)


def sweep(fn, values, *, name="x"):
    """Run fn(v) for each v and collect the results as a list of dicts.

    fn must return a dict. The sweep variable is added under `name`, so the
    result drops straight into to_excel().
    """
    rows = []
    for v in values:
        out = fn(v)
        if not isinstance(out, dict):
            raise TypeError("the swept function must return a dict of results")
        rows.append({name: v, **out})
    return rows


# -------------------------------------------------------------------- excel
def to_excel(path, sheets: dict, *, sources=None, summary=None, title=None):
    """Write a workbook that is genuinely usable.

    Every numeric cell is written as a number (not text), AutoFilter is on,
    the header row is frozen, and columns are sized to content. A Summary and
    a Sources sheet are always present, because a result you cannot trace is
    not an engineering deliverable.

        sheets  = {"Sweep": [ {...}, {...} ], ...}   list of dicts per sheet
        sources = [ ("what", "where it came from"), ... ]
        summary = [ ("quantity", value, "units"), ... ]
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color="FFFFFF", name="Calibri")
    head_fill = PatternFill("solid", fgColor="1F3864")

    def _write(ws, rows, headers=None):
        headers = headers or (list(rows[0].keys()) if rows else [])
        for j, hname in enumerate(headers, 1):
            c = ws.cell(row=1, column=j, value=hname)
            c.font, c.fill = head_font, head_fill
            c.alignment = Alignment(horizontal="left")
        for i, row in enumerate(rows, 2):
            for j, hname in enumerate(headers, 1):
                v = row.get(hname)
                # Numbers stay numbers. This is the single most common way a
                # delivered workbook turns out not to be filterable.
                if isinstance(v, bool):
                    v = str(v)
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    v = float(v) if isinstance(v, float) else v
                ws.cell(row=i, column=j, value=v)
        if rows:
            ws.auto_filter.ref = (f"A1:{get_column_letter(len(headers))}"
                                  f"{len(rows) + 1}")
        ws.freeze_panes = "A2"
        for j, hname in enumerate(headers, 1):
            width = max([len(str(hname))] +
                        [len(f"{r.get(hname)}") for r in rows[:200]]) + 3
            ws.column_dimensions[get_column_letter(j)].width = min(width, 42)

    # Summary first, so it is what opens.
    ws = wb.create_sheet("Summary")
    ws["A1"] = title or "AM5061 results"
    ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    r = 3
    for item in (summary or []):
        for j, v in enumerate(item, 1):
            ws.cell(row=r, column=j, value=v)
        r += 1
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    for sname, rows in sheets.items():
        _write(wb.create_sheet(sname[:31]), rows)

    ws = wb.create_sheet("Sources")
    _write(ws, [{"item": a, "source": b} for a, b in (sources or [])])

    wb.save(path)
    return path


# --------------------------------------------------------------- plot style
def style_plots():
    """Match the lecture decks, so figures in a report look like the slides."""
    import matplotlib as mpl
    mpl.rcParams.update({
        "figure.figsize": (7.2, 4.4), "figure.dpi": 110,
        "axes.edgecolor": MUTED, "axes.labelcolor": NAVY,
        "axes.titlecolor": NAVY, "axes.titlesize": 11.5,
        "axes.spines.top": False, "axes.spines.right": False,
        "axes.grid": True, "grid.alpha": 0.25, "grid.linewidth": 0.6,
        "xtick.color": MUTED, "ytick.color": MUTED,
        "font.size": 10, "legend.frameon": False,
        "axes.prop_cycle": mpl.cycler(color=[NAVY, ORANGE, BLUE, "#7F9DB9"]),
    })


# ------------------------------------------------- exchanger relations
def lmtd(dT1, dT2):
    """Log-mean temperature difference.

    Falls back to the arithmetic mean when the two ends are within 1% of each
    other, where the log form is numerically unstable and the two agree to
    better than 0.01% anyway.
    """
    dT1, dT2 = float(dT1), float(dT2)
    if dT1 <= 0 or dT2 <= 0:
        raise ValueError(
            f"temperature difference must be positive at both ends "
            f"(got {dT1:g} and {dT2:g}). A non-positive end means the streams "
            "cross, which no exchanger of this configuration can do."
        )
    if abs(dT1 - dT2) < 0.01*max(dT1, dT2):
        return 0.5*(dT1 + dT2)
    return (dT1 - dT2)/math.log(dT1/dT2)


def effectiveness(config, NTU, Cr):
    """Effectiveness for the standard configurations.

    config: 'counter', 'parallel', 'shell1'  (one shell pass, 2/4/... tube passes),
            'cross-both-unmixed' (approximate), 'cross-Cmax-mixed', 'cross-Cmin-mixed'

    Cr = C_min/C_max. Cr = 0 is the phase-change limit and every configuration
    collapses to the same expression, which is why boilers and condensers are
    easy and everything else is not.
    """
    if NTU < 0:
        raise ValueError("NTU cannot be negative")
    if not 0 <= Cr <= 1:
        raise ValueError(f"Cr must be between 0 and 1, got {Cr:g}")
    if Cr == 0:                       # phase change on one side
        return 1 - math.exp(-NTU)
    if config == "counter":
        if abs(Cr - 1) < 1e-12:
            return NTU/(1 + NTU)
        e = math.exp(-NTU*(1 - Cr))
        return (1 - e)/(1 - Cr*e)
    if config == "parallel":
        return (1 - math.exp(-NTU*(1 + Cr)))/(1 + Cr)
    if config == "shell1":
        r = math.sqrt(1 + Cr*Cr)
        e = math.exp(-NTU*r)
        return 2/(1 + Cr + r*(1 + e)/(1 - e))
    if config == "cross-both-unmixed":
        return 1 - math.exp((math.exp(-Cr*NTU**0.78) - 1)*NTU**0.22/Cr)
    if config == "cross-Cmax-mixed":
        return (1/Cr)*(1 - math.exp(-Cr*(1 - math.exp(-NTU))))
    if config == "cross-Cmin-mixed":
        return 1 - math.exp(-(1 - math.exp(-Cr*NTU))/Cr)
    raise ValueError(f"unknown configuration {config!r}")


def ntu_required(config, eps, Cr, hi=200.0):
    """Invert effectiveness() for NTU. Design direction, rather than rating."""
    eps_max = effectiveness(config, hi, Cr)
    if eps >= eps_max:
        raise ValueError(
            f"effectiveness {eps:g} is unreachable for {config} at Cr={Cr:g}; "
            f"the limit as NTU->infinity is {eps_max:.6f}. "
            "Change the configuration or accept less."
        )
    return solve(lambda n: effectiveness(config, n, Cr) - eps, 1.0,
                 bracket=(1e-9, hi))


# --------------------------------------------------------------- exergy
T0_REF, P0_REF = 303.15, 101325.0      # 30 C, sea level: the Chennai dead state


def exergy(st, T0=T0_REF, p0=P0_REF):
    """Specific flow exergy, J/kg:  (h - h0) - T0*(s - s0).

    The dead state is the ambient the plant actually sits in, so it is a
    DESIGN CHOICE, not a constant. Report which one you used - a Chennai
    dead state and a European one give different answers for the same plant.
    """
    ref = State(st.fluid, T=T0, P=p0)
    return (st.h - ref.h) - T0*(st.s - ref.s)
